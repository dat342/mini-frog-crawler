# -*- coding: utf-8 -*-
"""Web app quét URL - backend FastAPI."""
import asyncio
import csv
import io
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from crawler import CrawlJob, DEFAULT_PATTERNS, compute_duplicates

BASE_DIR = Path(__file__).parent
app = FastAPI(title="Mini Frog - Web URL Crawler")

_job: Optional[CrawlJob] = None
_task: Optional[asyncio.Task] = None


class CrawlRequest(BaseModel):
    mode: str = "auto"                 # auto | seeds
    url: str = ""
    seeds: list[str] = []
    max_urls: int = 1000
    patterns: dict = {}
    exclude: list = []
    respect_robots: bool = True
    concurrency: int = 5
    delay: float = 0.2


@app.get("/")
async def index():
    return FileResponse(BASE_DIR / "static" / "index.html")


@app.get("/api/defaults")
async def defaults():
    return {"patterns": DEFAULT_PATTERNS}


@app.post("/api/crawl")
async def start_crawl(req: CrawlRequest):
    global _job, _task
    if _job and _job.state == "running":
        raise HTTPException(409, "Đang có phiên crawl chạy. Bấm Dừng trước khi chạy mới.")
    try:
        _job = CrawlJob(req.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))
    _task = asyncio.create_task(_job.run())
    return {"ok": True}


@app.post("/api/stop")
async def stop_crawl():
    if _job:
        _job.stop()
    return {"ok": True}


@app.get("/api/status")
async def status():
    if not _job:
        return {"state": "idle", "fetched": 0, "queued": 0, "by_type": {}, "errors": 0,
                "elapsed": 0, "recent": [], "max_urls": 0, "message": ""}
    return _job.snapshot()


@app.get("/api/orphans")
async def orphans():
    if not _job:
        return {"sitemaps": [], "sitemap_count": 0, "orphans": [], "limited": False}
    return {
        "sitemaps": _job.sitemaps_found,
        "sitemap_count": len(_job.sitemap_urls),
        "orphans": _job.orphan_urls,
        "limited": len(_job.records) >= _job.max_urls,
    }


@app.get("/api/results")
async def results(since: int = 0):
    if not _job:
        return {"records": [], "total": 0}
    records = _job.records[since:]
    return {"records": records, "total": len(_job.records)}


# ---------------- Xuất file ----------------

URL_COLUMNS = [
    ("URL", "url", 55),
    ("Loại trang", "page_type_label", 18),
    ("Tầng 1", "_seg1", 20),
    ("Tầng 2", "_seg2", 22),
    ("Tầng 3", "_seg3", 22),
    ("Độ sâu thư mục", "folder_depth", 13),
    ("Status", "status", 8),
    ("Redirect", "_redirect", 30),
    ("Depth", "depth", 7),
    ("Title", "title", 45),
    ("Dài title", "title_length", 9),
    ("Meta description", "meta_description", 50),
    ("Dài meta", "meta_desc_length", 9),
    ("Số H1", "h1_count", 7),
    ("H1", "_h1", 40),
    ("Số H2", "h2_count", 7),
    ("Canonical", "canonical", 45),
    ("Canonical đúng", "_canonical_ok", 14),
    ("Meta robots", "meta_robots", 16),
    ("Ngày đăng", "date_published", 12),
    ("Ngày cập nhật", "date_modified", 13),
    ("Nguồn ngày", "date_source", 14),
    ("Thumbnail (og:image)", "og_image", 45),
    ("Số ảnh", "image_count", 7),
    ("Ảnh thiếu alt", "images_missing_alt", 12),
    ("Số từ", "word_count", 8),
    ("Inlinks", "inlink_count", 8),
    ("Anchor vào trang (mẫu)", "_anchors", 40),
    ("Tìm thấy tại", "found_on", 45),
    ("Phản hồi (ms)", "response_ms", 12),
    ("Lỗi", "error", 30),
]


def _row_for(record: dict):
    row = []
    for _, key, _w in URL_COLUMNS:
        if key.startswith("_seg"):
            idx = int(key[4:]) - 1
            segments = record.get("path_segments") or []
            row.append(segments[idx] if idx < len(segments) else "")
        elif key == "_anchors":
            anchors = []
            for s in record.get("inlink_sources") or []:
                if s["anchor"] and s["anchor"] not in anchors:
                    anchors.append(s["anchor"])
            row.append(" | ".join(anchors[:5]))
        elif key == "_redirect":
            row.append(" ; ".join(record.get("redirect_chain") or []))
        elif key == "_h1":
            row.append(" | ".join(record.get("h1_texts") or []))
        elif key == "_canonical_ok":
            ok = record.get("canonical_ok")
            row.append("" if record.get("status") != 200 else ("Đúng" if ok else "Sai/Thiếu"))
        else:
            value = record.get(key)
            row.append("" if value is None else value)
    return row


def _find_issues(records: list, duplicates: dict) -> list:
    """Danh sách vấn đề: (URL, vấn đề, chi tiết, tìm thấy tại)."""
    dup_titles = {u for urls in duplicates["titles"].values() for u in urls}
    dup_descs = {u for urls in duplicates["descriptions"].values() for u in urls}
    issues = []
    for r in records:
        url, found = r["url"], r.get("found_on", "")
        status = r.get("status")
        if not isinstance(status, int):
            issues.append((url, "Không truy cập được", r.get("error") or "", found))
            continue
        if status >= 400:
            issues.append((url, f"Lỗi HTTP {status}", "Link gãy" if status == 404 else "", found))
            continue
        if r.get("redirect_chain"):
            issues.append((url, "Redirect", " ; ".join(r["redirect_chain"]), found))
        if status != 200:
            continue
        if r.get("noindex"):
            issues.append((url, "Noindex", r.get("meta_robots") or "", found))
        if not r.get("title"):
            issues.append((url, "Thiếu title", "", found))
        elif r["title_length"] > 65:
            issues.append((url, "Title quá dài", f"{r['title_length']} ký tự (nên ≤ 65)", found))
        if not r.get("meta_description"):
            issues.append((url, "Thiếu meta description", "", found))
        elif r["meta_desc_length"] > 165:
            issues.append((url, "Meta quá dài", f"{r['meta_desc_length']} ký tự (nên ≤ 165)", found))
        if r.get("h1_count", 0) == 0:
            issues.append((url, "Thiếu H1", "", found))
        elif r["h1_count"] > 1:
            issues.append((url, "Nhiều H1", f"{r['h1_count']} thẻ H1", found))
        if r.get("canonical_ok") is False:
            detail = r.get("canonical") or "không có thẻ canonical"
            issues.append((url, "Canonical sai/thiếu", detail, found))
        if url in dup_titles:
            issues.append((url, "Title trùng lặp", r.get("title") or "", found))
        if url in dup_descs:
            issues.append((url, "Meta description trùng lặp", "", found))
        if r.get("images_missing_alt", 0) > 0:
            issues.append((url, "Ảnh thiếu alt", f"{r['images_missing_alt']}/{r['image_count']} ảnh", found))
        if r.get("word_count", 0) < 150:
            issues.append((url, "Thin content", f"{r.get('word_count', 0)} từ (nên ≥ 150)", found))
    return issues


HEADER_FILL = PatternFill("solid", fgColor="004AEF")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, widths=None):
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.freeze_panes = "A2"


@app.get("/api/export/xlsx")
async def export_xlsx():
    if not _job or not _job.records:
        raise HTTPException(400, "Chưa có dữ liệu để xuất")
    records = _job.records
    duplicates = compute_duplicates(records)
    issues = _find_issues(records, duplicates)

    wb = Workbook()

    # Sheet 1: Tổng quan
    ws = wb.active
    ws.title = "Tổng quan"
    snap = _job.snapshot()
    ws.append(["Chỉ số", "Giá trị"])
    ws.append(["Website", _job.seeds[0]])
    ws.append(["Tổng URL đã quét", len(records)])
    ws.append(["Thời gian crawl (giây)", snap["elapsed"]])
    for page_type, count in sorted(snap["by_type"].items()):
        from crawler import PAGE_TYPE_LABELS
        ws.append([f"Số trang: {PAGE_TYPE_LABELS.get(page_type, page_type)}", count])
    ws.append(["Lỗi truy cập / 4xx / 5xx", snap["errors"]])
    ws.append(["Nhóm title trùng lặp", len(duplicates["titles"])])
    ws.append(["Nhóm meta description trùng lặp", len(duplicates["descriptions"])])
    ws.append(["Trang thin content (< 150 từ)",
               sum(1 for r in records if r.get("status") == 200 and r.get("word_count", 0) < 150)])
    ws.append(["URL trong sitemap", len(_job.sitemap_urls)])
    ws.append(["Orphan pages (sitemap có, link nội bộ không)", len(_job.orphan_urls)])
    ws.append(["Tổng vấn đề phát hiện", len(issues)])
    _style_header(ws, [40, 60])

    # Sheet 2: URLs
    ws = wb.create_sheet("URLs")
    ws.append([c[0] for c in URL_COLUMNS])
    for r in records:
        ws.append(_row_for(r))
    _style_header(ws, [c[2] for c in URL_COLUMNS])

    # Sheet 3: Hình ảnh
    ws = wb.create_sheet("Hình ảnh")
    ws.append(["Trang", "URL ảnh", "Alt", "Thiếu alt"])
    for r in records:
        for img in r.get("images", []):
            ws.append([r["url"], img["src"], img["alt"], "Thiếu" if img["missing_alt"] else ""])
    _style_header(ws, [50, 60, 40, 10])

    # Sheet 4: Trùng lặp
    ws = wb.create_sheet("Trùng lặp")
    ws.append(["Loại", "Giá trị trùng", "Số trang", "Các URL"])
    for title, urls in duplicates["titles"].items():
        ws.append(["Title", title, len(urls), "\n".join(urls)])
    for desc, urls in duplicates["descriptions"].items():
        ws.append(["Meta description", desc, len(urls), "\n".join(urls)])
    _style_header(ws, [18, 50, 10, 70])

    # Sheet 5: Inlinks
    ws = wb.create_sheet("Inlinks")
    ws.append(["URL đích", "Tổng inlinks", "URL nguồn (mẫu)", "Anchor text"])
    for r in records:
        for s in (r.get("inlink_sources") or []):
            ws.append([r["url"], r.get("inlink_count", 0), s["from"], s["anchor"]])
    _style_header(ws, [55, 12, 55, 35])

    # Sheet 6: Orphan pages
    ws = wb.create_sheet("Orphan")
    ws.append(["URL trong sitemap nhưng không có link nội bộ trỏ tới"])
    for u in _job.orphan_urls:
        ws.append([u])
    _style_header(ws, [90])

    # Sheet 7: Vấn đề
    ws = wb.create_sheet("Vấn đề")
    ws.append(["URL", "Vấn đề", "Chi tiết", "Tìm thấy tại"])
    for row in issues:
        ws.append(list(row))
    _style_header(ws, [55, 26, 45, 50])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    host = _job.root_host.replace(".", "-")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="crawl-{host}.xlsx"'},
    )


@app.get("/api/export/csv")
async def export_csv():
    if not _job or not _job.records:
        raise HTTPException(400, "Chưa có dữ liệu để xuất")
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c[0] for c in URL_COLUMNS])
    for r in _job.records:
        writer.writerow(_row_for(r))
    data = "﻿" + buf.getvalue()  # BOM để Excel mở đúng tiếng Việt
    host = _job.root_host.replace(".", "-")
    return StreamingResponse(
        io.BytesIO(data.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="crawl-{host}.csv"'},
    )
